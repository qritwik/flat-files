from openpyxl import load_workbook

sku_code = []



#Path to the workbook
filepath = "ritwik.xlsx"
wb=load_workbook(filepath)
wb.create_sheet('sid1')

input_sheet = wb['Input']
output_sheet = wb['sid1']
rule_sheet = wb['FF']



for row in rule_sheet.iter_rows('A2:C14'):
	out = row[0].value
	inp = row[1].value
	rule = row[2].value

	#For copy --> 1
	if(rule == 1):
		for i in range(3,1000):
			if(input_sheet.cell(row = i, column = inp).value != None):
				output_sheet.cell(row = i, column = out).value = input_sheet.cell(row = i, column = inp).value

	#For concat --> 2
	elif(rule == 2):
		
		my_data = inp.split(',')
		
		for i in range(3,1000):
			for j in my_data:

				if(input_sheet.cell(row = i, column = int(j)).value != None):
					fin = ""

					for a in my_data:
						fin = fin +" "+str(input_sheet.cell(row = i, column = int(a)).value)
					output_sheet.cell(row = i, column = out).value = fin
				



wb.save('aaaaaaeer.xlsx')


		

			




	


    
    





